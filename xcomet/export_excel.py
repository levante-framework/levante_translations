#!/usr/bin/env python3
import argparse
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description='Export COMET per-segment scores to a color-coded Excel file.')
    p.add_argument('--lang', required=True, help='Target language code to use as translation column header (e.g., es-CO)')
    p.add_argument('--input_csv', type=Path, help='Path to segment_scores.csv (defaults to xcomet/output/<lang>/segment_scores.csv)')
    p.add_argument('--output_xlsx', type=Path, help='Output .xlsx file (defaults to xcomet/output/<lang>/segment_scores.xlsx)')
    p.add_argument('--sheet', default='Scores', help='Worksheet name')
    return p


def infer_paths(lang: str, input_csv: Optional[Path], output_xlsx: Optional[Path]) -> tuple[Path, Path]:
    base = Path(__file__).resolve().parent
    if input_csv is None:
        input_csv = base / 'output' / lang / 'segment_scores.csv'
    if output_xlsx is None:
        output_xlsx = base / 'output' / lang / 'segment_scores.xlsx'
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    return input_csv, output_xlsx


def auto_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 80)
        worksheet.column_dimensions[column].width = adjusted_width


def apply_header_and_alignment(ws):
    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # Right-align header cells except the first
    for col in range(2, ws.max_column + 1):
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='right')
    # Right-align all columns except the first (data rows)
    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            c = ws.cell(row=row, column=col)
            c.alignment = Alignment(horizontal='right')


def shrink_first_column_if_item_id(ws):
    # If first header is item_id, halve the width
    first_header = ws.cell(row=1, column=1).value
    if isinstance(first_header, str) and first_header.strip() == 'item_id':
        col_dim = ws.column_dimensions['A']
        try:
            current = float(col_dim.width or 10)
        except Exception:
            current = 10.0
        col_dim.width = max(6.0, current / 2.0)


def export_excel(lang: str, input_csv: Path, output_xlsx: Path, sheet_name: str):
    if not input_csv.exists():
        raise SystemExit(f'Input CSV not found: {input_csv}')

    df = pd.read_csv(input_csv)

    # Rename headers: source -> en, translation -> <lang>
    rename_map = {}
    if 'source' in df.columns:
        rename_map['source'] = 'en'
    if 'translation' in df.columns:
        rename_map['translation'] = lang
    df = df.rename(columns=rename_map)

    # Round score to 2 decimals
    if 'score' in df.columns:
        df['score'] = pd.to_numeric(df['score'], errors='coerce').round(2)

    # Write Excel
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # Determine score column index (1-based)
        headers = [cell.value for cell in ws[1]]
        try:
            score_col_idx = headers.index('score') + 1
        except ValueError:
            score_col_idx = None

        # Define fills
        fill_red = PatternFill(start_color='FFF8D7DA', end_color='FFF8D7DA', fill_type='solid')     # ~#F8D7DA
        fill_yellow = PatternFill(start_color='FFFFF3CD', end_color='FFFFF3CD', fill_type='solid')  # ~#FFF3CD
        fill_green = PatternFill(start_color='FFD4EDDA', end_color='FFD4EDDA', fill_type='solid')   # ~#D4EDDA

        # Color-code rows by score
        if score_col_idx is not None:
            for row in range(2, ws.max_row + 1):
                score_cell = ws.cell(row=row, column=score_col_idx)
                try:
                    score_val = float(score_cell.value) if score_cell.value is not None else None
                except Exception:
                    score_val = None
                if score_val is None:
                    continue
                if score_val < 0.30:
                    fill = fill_red
                elif score_val < 0.75:
                    fill = fill_yellow
                else:
                    fill = fill_green
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill

        apply_header_and_alignment(ws)
        auto_width(ws)
        shrink_first_column_if_item_id(ws)

    print(f'Wrote Excel: {output_xlsx}')


def main():
    args = build_parser().parse_args()
    input_csv, output_xlsx = infer_paths(args.lang, args.input_csv, args.output_xlsx)
    export_excel(args.lang, input_csv, output_xlsx, args.sheet)


if __name__ == '__main__':
    main()
