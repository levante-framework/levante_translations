#!/usr/bin/env python3
import argparse
import subprocess
import re
from pathlib import Path
from typing import List, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description='Run XCOMET/COMETKiwi for all languages and export a multi-sheet Excel with a Summary tab.')
    p.add_argument('--csv', required=True, help='Path or URL to Levante master CSV (e.g., translation_master.csv or https://...)')
    p.add_argument('--out_dir', type=Path, default=Path('xcomet/output'), help='Base output directory (default: xcomet/output)')
    p.add_argument('--langs', help='Comma-separated language codes to process (defaults to all language-like columns except en)')
    p.add_argument('--use_api', action='store_true', help='Use Python API (default).')
    p.add_argument('--use_cli', action='store_true', help='Use CLI instead of API')
    p.add_argument('--gpu', action='store_true', help='Use GPU if available (API) or pass --gpus 1 to CLI')
    p.add_argument('--matmul', choices=['medium', 'high'], help='If using GPU (API), set torch.set_float32_matmul_precision to this value')
    p.add_argument('--excel', type=Path, default=Path('xcomet/output/levante_comet_scores.xlsx'), help='Output Excel path')
    return p.parse_args()


def is_language_code(col: str) -> bool:
    # Match simple BCP-47 style: xx or xx-XX (allow extended like fr-CA)
    return re.fullmatch(r'^[a-z]{2}(-[A-Z]{2})?$', col) is not None


def fetch_remote_to_local(csv_loc: str, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    target = out_dir / '_remote_item_bank_translations.csv'
    try:
        import requests
        r = requests.get(csv_loc.strip(), timeout=60)
        r.raise_for_status()
        target.write_text(r.text, encoding='utf-8')
        print(f"Downloaded remote CSV to {target}")
        return target
    except Exception as e:
        raise SystemExit(f"Failed to download remote CSV: {e}")


def infer_languages(csv_loc: str) -> List[str]:
    # Read only headers row from URL or local path
    if '://' in csv_loc:
        df = pd.read_csv(csv_loc, nrows=1)
    else:
        df = pd.read_csv(Path(csv_loc), nrows=1)
    langs = []
    for col in df.columns:
        if col == 'en':
            continue
        if is_language_code(col):
            langs.append(col)
    return langs


def run_single_language(lang: str, csv_loc: str, out_dir: Path, use_api: bool, gpu: bool, use_cli: bool, matmul: str | None) -> Tuple[Path, Path]:
    lang_dir = out_dir / lang
    lang_dir.mkdir(parents=True, exist_ok=True)

    # Build command to run the per-language analysis
    cmd = ['python', str(Path(__file__).parent / 'run_xcomet.py'),
           '--lang', lang,
           '--csv', csv_loc,
           '--out_dir', str(out_dir)]

    if use_cli:
        cmd.append('--use_cli')
    else:
        cmd.append('--use_api')
        cmd.append('--allow_qe_fallback')
    if gpu:
        cmd.append('--gpu')
    if matmul:
        cmd.extend(['--matmul', matmul])

    print('Running:', ' '.join(cmd))
    subprocess.check_call(cmd)

    # Return paths to outputs
    return lang_dir / 'segment_scores.csv', lang_dir / 'segment_scores.md'


def color_code_worksheet(ws):
    headers = [cell.value for cell in ws[1]]
    try:
        score_col_idx = headers.index('score') + 1
    except ValueError:
        score_col_idx = None

    fill_red = PatternFill(start_color='FFF8D7DA', end_color='FFF8D7DA', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFFFF3CD', end_color='FFFFF3CD', fill_type='solid')
    fill_green = PatternFill(start_color='FFD4EDDA', end_color='FFD4EDDA', fill_type='solid')

    if score_col_idx is None:
        return

    for row in range(2, ws.max_row + 1):
        score_cell = ws.cell(row=row, column=score_col_idx)
        try:
            score_val = float(score_cell.value) if score_cell.value is not None else None
        except Exception:
            score_val = None
        if score_val is None:
            continue
        if score_val < 0.30:
            fill = fill_red
        elif score_val < 0.75:
            fill = fill_yellow
        else:
            fill = fill_green
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill


def auto_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                if val_len > max_length:
                    max_length = val_len
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 80)


def apply_header_and_alignment(ws):
    # Bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # Right-align header cells except the first
    for col in range(2, ws.max_column + 1):
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='right')
    # Right-align all columns except the first
    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')


def shrink_first_column_if_item_id(ws):
    # If first header is item_id, halve the width
    first_header = ws.cell(row=1, column=1).value
    if isinstance(first_header, str) and first_header.strip() == 'item_id':
        col_dim = ws.column_dimensions['A']
        try:
            current = float(col_dim.width or 10)
        except Exception:
            current = 10.0
        col_dim.width = max(6.0, current / 2.0)


def shrink_and_wrap_text_columns(ws, lang: str):
    # Identify columns by header names
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    targets = []
    if 'en' in headers:
        targets.append(headers.index('en') + 1)
    if lang in headers:
        targets.append(headers.index(lang) + 1)

    for col_idx in targets:
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        try:
            current = float(ws.column_dimensions[col_letter].width or 10)
        except Exception:
            current = 10.0
        ws.column_dimensions[col_letter].width = max(8.0, current / 2.0)
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal='right', wrap_text=True)
            ws.row_dimensions[row].height = 30


def build_summary_sheet(wb, summary_rows: List[dict]):
    if 'Summary' in wb.sheetnames:
        del wb['Summary']
    ws = wb.create_sheet('Summary', 0)
    headers = ['language', 'items', 'mean_score', '<0.30 (count)', '<0.75 (count)', '>=0.75 (count)', '<0.30 (%)', '<0.75 (%)', '>=0.75 (%)']
    ws.append(headers)
    for r in summary_rows:
        items = max(1, r['items'])
        ws.append([
            r['language'], r['items'], round(r['mean_score'], 2),
            r['lt_030'], r['lt_075'], r['ge_075'],
            f"{100.0 * r['lt_030']/items:.1f}%",
            f"{100.0 * r['lt_075']/items:.1f}%",
            f"{100.0 * r['ge_075']/items:.1f}%",
        ])
    apply_header_and_alignment(ws)
    auto_width(ws)
    shrink_first_column_if_item_id(ws)


def main():
    args = parse_args()
    csv_loc = args.csv.strip()
    # If URL, download once to local
    if '://' in csv_loc:
        csv_loc = str(fetch_remote_to_local(csv_loc, args.out_dir))

    langs = args.langs.split(',') if args.langs else infer_languages(csv_loc)
    if not langs:
        raise SystemExit('No language columns detected. Use --langs to specify, e.g., es-CO,de,fr-CA')

    # Collect per-language DataFrames and summary stats
    per_lang = []
    for lang in langs:
        csv_path, _ = run_single_language(lang, csv_loc, args.out_dir, use_api=args.use_api or not args.use_cli, gpu=args.gpu, use_cli=args.use_cli, matmul=args.matmul)
        df = pd.read_csv(csv_path)
        # Rename columns
        rename_map = {}
        if 'source' in df.columns:
            rename_map['source'] = 'en'
        if 'translation' in df.columns:
            rename_map['translation'] = lang
        df = df.rename(columns=rename_map)
        # Round score
        if 'score' in df.columns:
            df['score'] = pd.to_numeric(df['score'], errors='coerce').round(2)
        # Reorder columns: item_id, score, en, <lang>, then rest
        cols = list(df.columns)
        desired = ['item_id']
        if 'score' in cols:
            desired.append('score')
        if 'en' in cols:
            desired.append('en')
        if lang in cols:
            desired.append(lang)
        desired += [c for c in cols if c not in desired]
        df = df[desired]

        # Compute summary stats
        scores = pd.to_numeric(df['score'], errors='coerce') if 'score' in df.columns else pd.Series(dtype=float)
        items = int(scores.notna().sum())
        mean_score = float(scores.mean()) if items > 0 else 0.0
        lt_030 = int((scores < 0.30).sum()) if items > 0 else 0
        lt_075 = int(((scores >= 0.30) & (scores < 0.75)).sum()) if items > 0 else 0
        ge_075 = int((scores >= 0.75).sum()) if items > 0 else 0
        per_lang.append({'language': lang, 'df': df, 'items': items, 'mean_score': mean_score, 'lt_030': lt_030, 'lt_075': lt_075, 'ge_075': ge_075})

    # Write all sheets
    args.excel.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(args.excel, engine='openpyxl') as writer:
        for entry in per_lang:
            ws_name = entry['language'][:31]
            entry['df'].to_excel(writer, index=False, sheet_name=ws_name)

    # Post-process: load workbook, add Summary sheet, color and auto-width
    wb = load_workbook(args.excel)
    build_summary_sheet(wb, per_lang)
    for entry in per_lang:
        ws = wb[entry['language'][:31]]
        apply_header_and_alignment(ws)
        color_code_worksheet(ws)
        auto_width(ws)
        shrink_first_column_if_item_id(ws)
        shrink_and_wrap_text_columns(ws, entry['language'])
    wb.save(args.excel)
    print(f'Wrote multi-sheet Excel: {args.excel}')


if __name__ == '__main__':
    main()
